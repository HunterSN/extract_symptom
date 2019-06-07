import openpyxl
wb = openpyxl.load_workbook('temp1.xlsx')
hhh = 1-2

sheet = wb.active

for hang in range(1, sheet.max_row + 1):
        fc = sheet.cell(row = hang, column = 2 ).value
        sc = sheet.cell(row = hang, column = 3 ).value
        dabianku = ['量可','稀','质干','费力','不尽感','黏','质可','不成形']
        binglieku = ['、','及','偶有']
        xiaobianku = ['浑浊','色黄','淋漓','泡沫','絮状','异味','不畅']
        yeniaoku = ['夜尿'] ##第一列
        tanku = ['咳出','白','不易咳出','难咳','血丝'] ##第二列
        yinshuiku = ['欲']
        bianmiku = ['未行','开塞露','日','助']
        diyilieku = ['吸氧']
        dierlieku = ['性']
        fuzaku = ['屈伸','跛行']
    
        dabian = '大便'
        xiaobian = '小便'
        yeniao = '夜尿频'
        tan = '痰'
        yinshui = '饮水'
        bianmi = '便秘'

        for a1 in tanku:
            zzz = str(sc)
            
            aa1 = zzz.find(a1)
            if aa1 != hhh:
                sheet.cell(row = hang, column =4).value = tan + fc
                break
                pass
            else:
                #sheet.cell(row = hang, column =4).value = "D"
            
                pass
            pass  
        pass 
pass
wb.save('temp1.xlsx')
#wb.close("temp.xlsx")
