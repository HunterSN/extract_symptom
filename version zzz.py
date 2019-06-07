import openpyxl
wb = openpyxl.load_workbook('100001-200026.xlsx')
sheet = wb.active
hhh = 1-2


for hang in range(1, sheet.max_row + 1):
    fc = sheet.cell(row = hang, column = 2 ).value
    sc = sheet.cell(row = hang, column = 3 ).value
    dabianku = ['量可','通便','稀','质干','费力','不尽感','黏','质可','不成形','不畅']
    binglieku = ['、','及','偶有']
    xiaobianku = ['浑浊','色黄','淋漓','泡沫','絮状','异味']
    yeniaoku = ['夜尿'] ##第一列
    tanku = ['咳出','白','不易咳出','难咳','血丝'] ##第二列
    yinshuiku = ['欲']
    bianmiku = ['未行','开塞露','日','助']
    diyilieku = ['吸氧']
    dierlieku = ['性']
    fuzaku = ['屈伸','跛行']
    
    dabian = '大便'
    xiaobian = '小便'
    yeniao = '夜尿'
    tan = '痰'
    yinshui = '饮水'
    bianmi = '便秘'


    for a1 in dabianku:
        zzz = str(fc)
        aa1 = zzz.find(a1)
        if aa1 != hhh:
            sheet.cell(row = hang, column =4).value = dabian + fc
            break
            pass
        else:
            for b1 in binglieku:
                bb1 = fc.find(b1)
                if bb1 != hhh:
                    sheet.cell(row = hang, column =4).value = fc
                    break
                    pass
                else:
                    for c1 in xiaobianku:
                        cc1 = fc.find(c1)
                        if cc1 != hhh:
                            sheet.cell(row = hang, column =4).value =xiaobian + fc
                            break
                            pass
                        else:
                            for d1 in yeniaoku:
                                dd1 = fc.find(d1)
                                if dd1 != hhh:
                                    sheet.cell(row = hang, column =4).value =yeniao + fc
                                    break
                                    pass
                                else:
                                    for e1 in tanku:
                                        ee1 = fc.find(e1)
                                        if ee1 != hhh:
                                            sheet.cell(row = hang, column =4).value =tan + fc
                                            break
                                            pass
                                        else:
                                            for ff1 in binglieku:
                                                fff1 = fc.find(ff1)
                                                if fff1 != hhh:
                                                    sheet.cell(row = hang, column =4).value =fc + yinshui
                                                    break
                                                    pass
                                                else:
                                                    for g1 in binglieku:
                                                        gg1 = fc.find(g1)
                                                        if gg1 != hhh:
                                                            sheet.cell(row = hang, column =4).value = bianmi
                                                            break
                                                            pass
                                                        else:
                                                            for h1 in binglieku:
                                                                hh1 = fc.find(h1)
                                                                if hh1 != hhh:
                                                                    sheet.cell(row = hang, column =4).value = fc
                                                                    break
                                                                    pass
                                                                else:
                                                                    for i1 in binglieku:
                                                                        ii1 = fc.find(i1)
                                                                        if ii1 != hhh:
                                                                            sheet.cell(row = hang, column =4).value = sc
                                                                            break
                                                                            pass
                                                                        else:
                                                                            for j1 in binglieku:
                                                                                jj1 = fc.find(j1)
                                                                                if jj1 != hhh:
                                                                                    sheet.cell(row = hang, column =4).value = sheet.cell(row = hang - 1, column =4).value + fc
                                                                                    break
                                                                                    pass
                                                                                else:
                                                                                    sheet.cell(row = hang, column =4).value = "D"           
                                                                                    pass
                                                                                pass             
                                                                            pass
                                                                        pass            
                                                                    pass
                                                                pass            
                                                            pass
                                                        pass        
                                                    pass
                                                pass             
                                            pass
                                        pass            
                                    pass
                                pass 
                            pass
                        pass             
                    pass
                pass            
            pass
        pass    
pass
wb.save('temp1.xlsx')

sheet.cell(row = hang, column =4).value = "D"

